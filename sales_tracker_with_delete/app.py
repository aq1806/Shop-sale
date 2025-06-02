from flask import Flask, render_template, request, redirect, url_for
import openpyxl
from datetime import datetime

app = Flask(__name__)

EXCEL_FILE = 'sales.xlsx'

def init_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        # Check if header exists, if not create it
        if ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None:
            ws.append(['ID', 'Date', 'Amount', 'Items', 'PaymentType'])
            wb.save(EXCEL_FILE)
        elif ws.max_column < 5:
            # Add PaymentType column if missing
            ws.cell(row=1, column=5, value='PaymentType')
            wb.save(EXCEL_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['ID', 'Date', 'Amount', 'Items', 'PaymentType'])
        wb.save(EXCEL_FILE)

def get_all_sales():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    sales = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        sales.append(row)  # (ID, Date, Amount, Items, PaymentType)
    return sales

def get_sales_by_date(date_str):
    all_sales = get_all_sales()
    filtered = [s for s in all_sales if s[1] == date_str]
    return filtered

def get_next_id():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        try:
            val_int = int(val)
            if val_int > max_id:
                max_id = val_int
        except (TypeError, ValueError):
            continue
    return max_id + 1

@app.route('/', methods=['GET', 'POST'])
def index():
    init_excel()
    today_str = datetime.now().strftime('%Y-%m-%d')

    if request.method == 'POST':
        amount = request.form.get('amount')
        items_list = request.form.getlist('items')
        items_list = [item.strip() for item in items_list if item.strip()]
        items_str = ' + '.join(items_list)
        payment_type = request.form.get('payment_type', 'Cash')

        if amount and items_str:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            sale_id = get_next_id()
            ws.append([sale_id, today_str, round(float(amount), 3), items_str, payment_type])
            wb.save(EXCEL_FILE)
            return redirect(url_for('index'))

    sales = get_sales_by_date(today_str)
    total = sum(float(s[2]) for s in sales)
    count = len(sales)

    return render_template('index.html', sales=sales, total=total, count=count)

@app.route('/previous', methods=['GET', 'POST'])
def previous():
    init_excel()

    selected_date = None
    sales = []
    total = 0
    count = 0

    if request.method == 'POST':
        selected_date = request.form.get('date')
        if selected_date:
            sales = get_sales_by_date(selected_date)
            total = sum(float(s[2]) for s in sales)
            count = len(sales)

    return render_template('previous.html', sales=sales, total=total, count=count, selected_date=selected_date or '')

@app.route('/delete/<int:row_id>', methods=['POST'])
def delete(row_id):
    from_page = request.args.get('from_page', 'index')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    row_to_delete = None
    for row in ws.iter_rows(min_row=2):
        cell = row[0]
        if cell.value == row_id:
            row_to_delete = cell.row
            break

    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(EXCEL_FILE)

    if from_page == 'previous':
        return redirect(url_for('previous'))
    else:
        return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

